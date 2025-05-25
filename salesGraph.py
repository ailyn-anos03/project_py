import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import os
from datetime import datetime
from subprocess import call


def open_py_file():
    root.destroy()
    


def load_data_and_plot():
    file_path = "inventory.xlsx"

    if not os.path.exists(file_path):
        messagebox.showerror("Error", f"File not found: {file_path}")
        return

    try:
        wb = load_workbook(file_path, data_only=True)
        if "History" not in wb.sheetnames:
            messagebox.showerror("Error", "Worksheet 'History' not found.")
            return
        ws = wb["History"]
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load Excel file:\n{e}")
        return

    quantities = []
    timestamps = []
    last_valid_timestamp = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        quantity = row[2] if len(row) > 2 else None
        timestamp = row[4] if len(row) > 4 else None

        if timestamp is None and last_valid_timestamp is not None:
            timestamp = last_valid_timestamp
        elif timestamp is not None:
            last_valid_timestamp = timestamp

        if quantity is not None and timestamp is not None:
            try:
                quantities.append(float(quantity))
                timestamps.append(str(timestamp))
            except:
                continue

    if len(quantities) < 2:
        messagebox.showwarning("Data Error", "Not enough valid data to plot.")
        return

    draw_plot(timestamps, quantities)

def draw_plot(labels, values):
    canvas.delete("all")
    width = 1100
    height = 500
    margin = 100

    max_val = max(values)
    min_val = min(values)
    range_val = max_val - min_val or 1

    n = len(values)
    step_x = (width - 2 * margin) / (n - 1)
    scale_y = (height - 2 * margin) / range_val

    points = []
    for i, val in enumerate(values):
        x = margin + i * step_x
        y = height - margin - ((val - min_val) * scale_y)
        points.append((x, y))

    # Draw axes
    canvas.create_line(margin, margin, margin, height - margin, width=2, fill="white")
    canvas.create_line(margin, height - margin, width - margin, height - margin, width=2, fill="white")

    # Draw thicker lines (removed dots)
    for i in range(len(points) - 1):
        canvas.create_line(*points[i], *points[i + 1], fill="cyan", width=4)

    # Y-axis labels (5 steps)
    for i in range(6):
        val = min_val + i * (range_val / 5)
        y = height - margin - (val - min_val) * scale_y
        canvas.create_text(margin - 15, y, text=f"{val:.1f}", anchor="e", font=("Arial", 11), fill="white")

    # X-axis labels (max 6 visible)
    label_interval = max(1, len(labels) // 6)
    for i in range(0, len(points), label_interval):
        x, _ = points[i]
        try:
            dt = datetime.fromisoformat(labels[i])
            label = dt.strftime("%Y-%m-%d")
        except:
            label = labels[i][:10]
        canvas.create_text(x, height - margin + 40, text=label, anchor="n", angle=30, font=("Arial", 10), fill="white")

    # # Axis titles
    canvas.create_text(width / 2, height - margin + 100, text="Timestamp", font=("Arial", 12, "bold"), fill="white")
    canvas.create_text(margin - 80, height / 2, text="Quantity", font=("Arial", 12, "bold"), angle=90, fill="white")

# === GUI Setup ===


root = tk.Tk()
root.title("Inventory History Graph")
root.geometry("1100x650")
root.resizable(False, False)
root.configure(bg= "black")

graph_frame = tk.Frame(root)
graph_frame.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="nsew")

canvas = tk.Canvas(graph_frame, width=1500, height=650, bg="black", highlightthickness=0)
canvas.grid(row=0, column=0, sticky="nsew")

load_button = ctk.CTkButton(root, text="    Exit    ", command=open_py_file, font=("Arial", 14))

load_button.grid(row=1, column=0, pady=(10, 20))

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

load_data_and_plot()

root.mainloop()
